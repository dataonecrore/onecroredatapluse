"""Resumable CSV importer for large customer data sets.

This command connects directly to Postgres and uses COPY into a temporary table.
It is deliberately separate from the web API: never upload a 10-million-row file
through the browser or expose the database password to the frontend.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import sys
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, TextIO


VALID_DUPLICATE_MODES = ("preserve", "skip-phone")
PHONE_DIGITS = re.compile(r"\D+")
WHITESPACE = re.compile(r"\s+")
HEADER_ALIASES = {
    "adress": "address",
    "customer adress": "customer address",
    "address line": "address",
    "address line 1": "address",
}


@dataclass(frozen=True)
class CustomerRow:
    source_row: int
    customer_name: str
    phone: str | None
    address: str | None
    source_customer_id: str | None


@dataclass(frozen=True)
class RejectedRow:
    source_row: int
    error_code: str
    error_message: str


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def normalized_phone(value: str) -> str:
    return PHONE_DIGITS.sub("", value or "")


def clean_text(value: object | None) -> str:
    return WHITESPACE.sub(" ", str(value or "")).strip()


def validate_headers(fieldnames: list[str] | None, columns: dict[str, str | None]) -> None:
    if not fieldnames:
        raise ValueError("CSV file has no header row")
    optional_columns = {"phone", "house_no", "street", "village", "post_office"}
    missing = [
        column
        for key, column in columns.items()
        if column and key not in optional_columns and column not in fieldnames
    ]
    if missing:
        raise ValueError(f"CSV is missing configured column(s): {', '.join(missing)}")


def canonicalize_headers(
    fieldnames: list[object] | tuple[object, ...] | None,
    columns: dict[str, str | None],
) -> list[object] | None:
    if fieldnames is None:
        return None
    configured = {
        re.sub(r"[^a-z0-9]+", " ", str(column).lower()).strip(): column
        for column in columns.values()
        if column
    }
    canonical = []
    for fieldname in fieldnames:
        key = re.sub(r"[^a-z0-9]+", " ", str(fieldname).lower()).strip()
        key = HEADER_ALIASES.get(key, key)
        canonical.append(configured.get(key, fieldname))
    return canonical


def parse_customer_row(
    row: dict[str, str | None], source_row: int, columns: dict[str, str | None]
) -> CustomerRow | RejectedRow:
    name = clean_text(row.get(columns["name"] or ""))
    phone = clean_text(row.get(columns["phone"] or "")) or None
    if not name:
        return RejectedRow(source_row, "missing_name", "Customer name is required")
    if phone and len(normalized_phone(phone)) < 3:
        return RejectedRow(
            source_row,
            "invalid_phone",
            "Phone must contain at least 3 digits",
        )
    address_columns = (
        columns.get("address"),
        columns.get("house_no"),
        columns.get("street"),
        columns.get("village"),
        columns.get("post_office"),
        columns.get("city"),
        columns.get("state"),
        columns.get("pin_code"),
    )
    source_id_column = columns.get("source_id")
    address_parts = [clean_text(row.get(column)) for column in address_columns if column]
    address = ", ".join(part for part in address_parts if part) or None
    source_id = clean_text(row.get(source_id_column)) if source_id_column else None
    source_id = source_id or None
    return CustomerRow(source_row, name, phone, address, source_id)


def iter_batches(
    source: TextIO,
    columns: dict[str, str | None],
    batch_size: int,
    after_source_row: int = 0,
) -> Iterator[tuple[list[CustomerRow], list[RejectedRow], int]]:
    reader = csv.DictReader(source)
    reader.fieldnames = canonicalize_headers(reader.fieldnames, columns)
    validate_headers(reader.fieldnames, columns)
    accepted: list[CustomerRow] = []
    rejected: list[RejectedRow] = []
    batch_end = after_source_row

    for source_row, row in enumerate(reader, start=1):
        if source_row <= after_source_row:
            continue
        parsed = parse_customer_row(row, source_row, columns)
        if isinstance(parsed, CustomerRow):
            accepted.append(parsed)
        else:
            rejected.append(parsed)
        batch_end = source_row
        if len(accepted) + len(rejected) >= batch_size:
            yield accepted, rejected, batch_end
            accepted, rejected = [], []

    if accepted or rejected:
        yield accepted, rejected, batch_end


def iter_file_batches(
    source_path: Path,
    columns: dict[str, str | None],
    batch_size: int,
    encoding: str = "utf-8-sig",
    after_source_row: int = 0,
) -> Iterator[tuple[list[CustomerRow], list[RejectedRow], int]]:
    """Stream CSV or Excel rows without loading the complete file."""
    extension = source_path.suffix.lower()
    if extension == ".csv":
        with source_path.open("r", encoding=encoding, newline="") as source:
            yield from iter_batches(source, columns, batch_size, after_source_row)
        return

    if extension == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(source_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
    elif extension == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(source_path, on_demand=True)
        sheet = workbook.sheet_by_index(0)
        rows = (sheet.row_values(row_index) for row_index in range(sheet.nrows))
    else:
        raise ValueError("Supported production files are CSV, XLS, and XLSX")

    try:
        headers = canonicalize_headers(next(rows, None), columns)
        validate_headers(headers, columns)
        accepted: list[CustomerRow] = []
        rejected: list[RejectedRow] = []
        batch_end = after_source_row
        for source_row, values in enumerate(rows, start=1):
            if source_row <= after_source_row:
                continue
            parsed = parse_customer_row(dict(zip(headers, values)), source_row, columns)
            if isinstance(parsed, CustomerRow):
                accepted.append(parsed)
            else:
                rejected.append(parsed)
            batch_end = source_row
            if len(accepted) + len(rejected) >= batch_size:
                yield accepted, rejected, batch_end
                accepted, rejected = [], []
        if accepted or rejected:
            yield accepted, rejected, batch_end
    finally:
        if extension == ".xlsx":
            workbook.close()
        else:
            workbook.release_resources()


def _create_temp_staging(cursor) -> None:
    cursor.execute(
        """
        create temporary table if not exists customer_import_stage (
          source_row bigint not null,
          customer_name text not null,
          phone text,
          address text,
          source_customer_id text
        ) on commit delete rows
        """
    )


def _copy_rows(cursor, rows: list[CustomerRow]) -> None:
    if not rows:
        return
    with cursor.copy(
        "copy customer_import_stage "
        "(source_row, customer_name, phone, address, source_customer_id) from stdin"
    ) as copy:
        for row in rows:
            copy.write_row(
                (
                    row.source_row,
                    row.customer_name,
                    row.phone,
                    row.address,
                    row.source_customer_id,
                )
            )


def _insert_staged(cursor, job_id: uuid.UUID, duplicate_mode: str) -> int:
    if duplicate_mode == "preserve":
        selection = "select * from customer_import_stage"
    else:
        selection = """
          select *
          from (
            select customer_import_stage.*,
                   nullif(regexp_replace(coalesce(phone, ''), '\\D', '', 'g'), '')
                     as normalized_phone,
                   row_number() over (
                     partition by nullif(
                       regexp_replace(coalesce(phone, ''), '\\D', '', 'g'), ''
                     )
                     order by source_row
                   ) as duplicate_rank
            from customer_import_stage
          ) ranked
          where normalized_phone is null or duplicate_rank = 1
        """

    duplicate_filter = ""
    if duplicate_mode == "skip-phone":
        duplicate_filter = """
          where stage.normalized_phone is null or not exists (
            select 1 from public.customers existing
            where existing.normalized_phone = stage.normalized_phone
          )
        """

    cursor.execute(
        f"""
        with stage as ({selection}), inserted as (
          insert into public.customers (
            name, phone, address, source_customer_id,
            import_job_id, import_source_row
          )
          select stage.customer_name, stage.phone, stage.address,
                 stage.source_customer_id, %s, stage.source_row
          from stage
          {duplicate_filter}
          on conflict (import_job_id, import_source_row)
            where import_job_id is not null and import_source_row is not null
          do nothing
          returning 1
        )
        select count(*) from inserted
        """,
        (job_id,),
    )
    return int(cursor.fetchone()[0])


def _load_or_create_job(
    cursor,
    job_id: uuid.UUID | None,
    source_path: Path,
    source_hash: str,
    duplicate_mode: str,
):
    if job_id:
        cursor.execute(
            """
            select id, source_sha256, duplicate_mode, status, last_committed_row,
                   rows_inserted, rows_skipped, rows_rejected
            from private.customer_import_jobs where id = %s
            """,
            (job_id,),
        )
        job = cursor.fetchone()
        if not job:
            raise ValueError(f"Import job {job_id} does not exist")
        if job[1] != source_hash or job[2] != duplicate_mode:
            raise ValueError("Resume file checksum or duplicate mode does not match the job")
        return job

    cursor.execute(
        """
        select id, source_sha256, duplicate_mode, status, last_committed_row,
               rows_inserted, rows_skipped, rows_rejected
        from private.customer_import_jobs
        where source_sha256 = %s and duplicate_mode = %s
        """,
        (source_hash, duplicate_mode),
    )
    existing = cursor.fetchone()
    if existing:
        return existing

    new_id = uuid.uuid4()
    cursor.execute(
        """
        insert into private.customer_import_jobs
          (id, source_filename, source_sha256, duplicate_mode)
        values (%s, %s, %s, %s)
        returning id, source_sha256, duplicate_mode, status, last_committed_row,
                  rows_inserted, rows_skipped, rows_rejected
        """,
        (new_id, source_path.name, source_hash, duplicate_mode),
    )
    return cursor.fetchone()


def import_csv(args: argparse.Namespace) -> uuid.UUID:
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("Install backend requirements before running the importer") from exc

    source_path = Path(args.file).expanduser().resolve(strict=True)
    if source_path.suffix.lower() not in {".csv", ".xls", ".xlsx"}:
        raise ValueError("Production bulk import supports CSV, XLS, and XLSX files")
    source_hash = file_sha256(source_path)
    requested_job_id = uuid.UUID(args.job_id) if args.job_id else None
    columns = {
        "name": args.name_column,
        "phone": args.phone_column,
        "address": args.address_column,
        "house_no": args.house_no_column,
        "street": args.street_column,
        "village": args.village_column,
        "post_office": args.post_office_column,
        "city": args.city_column,
        "state": args.state_column,
        "pin_code": args.pin_code_column,
        "source_id": args.source_id_column,
    }

    with psycopg.connect(args.database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute("select pg_advisory_lock(hashtextextended(%s, 0))", (source_hash,))
            job = _load_or_create_job(
                cursor, requested_job_id, source_path, source_hash, args.duplicate_mode
            )
            job_id = job[0]
            if job[3] == "completed":
                print(f"Import job {job_id} is already completed")
                return job_id
            connection.commit()

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    update private.customer_import_jobs
                    set status = 'running', started_at = coalesce(started_at, now()),
                        error_message = null, updated_at = now()
                    where id = %s
                    """,
                    (job_id,),
                )
                connection.commit()

            last_row = int(job[4])
            for accepted, rejected, batch_end in iter_file_batches(
                source_path, columns, args.batch_size, args.encoding, last_row
            ):
                    with connection.cursor() as cursor:
                        _create_temp_staging(cursor)
                        _copy_rows(cursor, accepted)
                        inserted = _insert_staged(cursor, job_id, args.duplicate_mode)
                        skipped = len(accepted) - inserted
                        if rejected:
                            cursor.executemany(
                                """
                                insert into private.customer_import_rejections
                                  (import_job_id, source_row, error_code, error_message)
                                values (%s, %s, %s, %s)
                                on conflict (import_job_id, source_row) do nothing
                                """,
                                [
                                    (job_id, row.source_row, row.error_code, row.error_message)
                                    for row in rejected
                                ],
                            )
                        cursor.execute(
                            """
                            update private.customer_import_jobs
                            set last_committed_row = %s,
                                rows_inserted = rows_inserted + %s,
                                rows_skipped = rows_skipped + %s,
                                rows_rejected = rows_rejected + %s,
                                updated_at = now()
                            where id = %s
                            """,
                            (batch_end, inserted, skipped, len(rejected), job_id),
                        )
                    connection.commit()
                    print(
                        f"job={job_id} committed_through_row={batch_end} "
                        f"inserted={inserted} skipped={skipped} rejected={len(rejected)}",
                        flush=True,
                    )

            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    update private.customer_import_jobs
                    set status = 'completed', completed_at = now(), updated_at = now()
                    where id = %s
                    """,
                    (job_id,),
                )
            connection.commit()
            return job_id
        except Exception as exc:
            connection.rollback()
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    update private.customer_import_jobs
                    set status = 'failed', error_message = %s, updated_at = now()
                    where id = %s
                    """,
                    (str(exc)[:1000], job_id),
                )
            connection.commit()
            raise
        finally:
            with connection.cursor() as cursor:
                cursor.execute("select pg_advisory_unlock(hashtextextended(%s, 0))", (source_hash,))
            connection.commit()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Resumable customer CSV bulk importer")
    parser.add_argument("--file", required=True, help="Path to the source CSV")
    parser.add_argument(
        "--database-url",
        default=os.getenv("SUPABASE_DB_URL"),
        help="Direct/session-mode Postgres URL (or SUPABASE_DB_URL)",
    )
    parser.add_argument("--duplicate-mode", required=True, choices=VALID_DUPLICATE_MODES)
    parser.add_argument("--name-column", default="Customer Name")
    parser.add_argument("--phone-column", default="Customer Phone")
    parser.add_argument("--address-column", default="Customer Address")
    parser.add_argument("--house-no-column", default="House No")
    parser.add_argument("--street-column", default="Street")
    parser.add_argument("--village-column", default="Village")
    parser.add_argument("--post-office-column", default="Post office")
    parser.add_argument("--city-column", default="City")
    parser.add_argument("--state-column", default="State")
    parser.add_argument("--pin-code-column", default="PIN Code")
    parser.add_argument("--source-id-column")
    parser.add_argument("--job-id", help="Existing job UUID to resume")
    parser.add_argument("--batch-size", type=int, default=50_000)
    parser.add_argument("--encoding", default="utf-8-sig")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if not args.database_url:
        parser.error("--database-url or SUPABASE_DB_URL is required")
    if args.batch_size < 1 or args.batch_size > 250_000:
        parser.error("--batch-size must be between 1 and 250000")
    try:
        job_id = import_csv(args)
    except Exception as exc:
        print(f"Import failed: {exc}", file=sys.stderr)
        return 1
    print(f"Import completed: job={job_id}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
