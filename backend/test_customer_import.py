import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook


os.environ.setdefault("SUPABASE_URL", "https://example.supabase.co")
os.environ.setdefault("SUPABASE_KEY", "test-service-key")

from backend import main


class CustomerImportTests(unittest.TestCase):
    def test_standard_import_uses_larger_processing_and_write_batches(self):
        self.assertEqual(main.IMPORT_ROW_BATCH_SIZE, 5_000)
        self.assertEqual(main.IMPORT_WRITE_BATCH_SIZE, 500)
        self.assertEqual(main.IMPORT_LOOKUP_BATCH_SIZE, 250)

    def test_import_threshold_defaults_to_one_hundred_megabytes(self):
        self.assertEqual(main.SMALL_IMPORT_SIZE, 100 * 1024 * 1024)

    def test_thirty_two_megabyte_import_uses_standard_importer(self):
        self.assertLess(32 * 1024 * 1024, main.SMALL_IMPORT_SIZE)

    def test_import_customers_have_matching_bulk_payload_keys(self):
        first = main._build_import_customer(
            {"name": "First", "phone": "111"}
        )
        second = main._build_import_customer(
            {
                "name": "Second",
                "phone": "222",
                "email": "second@example.com",
                "address": "Pune",
            }
        )

        self.assertEqual(set(first), set(second))
        self.assertIsNone(first["email"])
        self.assertEqual(first["status"], "Active")
        self.assertEqual(second["email"], "second@example.com")

    def test_import_row_issues_distinguish_blank_and_incomplete_rows(self):
        self.assertEqual(main.get_import_row_issue({}), "blank")
        self.assertEqual(
            main.get_import_row_issue({"name": "", "phone": ""}),
            "blank",
        )
        self.assertIsNone(main.get_import_row_issue({"name": "Aarav", "phone": ""}))
        self.assertEqual(
            main.get_import_row_issue({"name": "", "phone": "9000000001"}),
            "missing_name",
        )
        self.assertEqual(
            main.get_import_row_issue({"name": "", "phone": "", "notes": "Review"}),
            "missing_name",
        )
        self.assertIsNone(
            main.get_import_row_issue({"name": "Aarav", "phone": "9000000001"})
        )

    @patch("backend.main._create_customers_batch")
    @patch("backend.main._find_duplicates_batch", return_value={})
    @patch("backend.main.iter_import_rows")
    def test_small_import_ignores_blank_rows_and_reports_invalid_reasons(
        self, iter_rows, _find_duplicates, create_batch
    ):
        iter_rows.return_value = iter(
            [
                {"name": "Aarav", "phone": "9000000001"},
                {},
                {"name": "", "phone": ""},
                {"name": "Priya", "phone": ""},
                {"name": "", "phone": "9000000002"},
            ]
        )
        job_id = "blank-row-test"
        main.import_jobs[job_id] = {}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as source:
            source_path = Path(source.name)

        main.import_customers(job_id, source_path, ".xlsx", "new", "phone")

        job = main.import_jobs.pop(job_id)
        self.assertEqual(job["status"], "ready")
        self.assertEqual(job["processed"], 3)
        self.assertEqual(job["blank_rows_ignored"], 2)
        self.assertEqual(job["invalid"], 1)
        self.assertEqual(
            job["invalid_reasons"],
            {"missing_name": 1},
        )
        create_batch.assert_called_once()
        self.assertEqual(len(create_batch.call_args.args[1]), 2)

    @patch("backend.main._create_customers_batch")
    @patch("backend.main._find_duplicates_batch", return_value={})
    @patch("backend.main.iter_import_rows")
    def test_standard_import_processes_five_thousand_rows_per_cycle(
        self, iter_rows, find_duplicates, create_batch
    ):
        iter_rows.return_value = iter(
            {
                "name": f"Customer {index}",
                "phone": str(9_000_000_000 + index),
            }
            for index in range(5_001)
        )
        job_id = "five-thousand-row-cycle-test"
        main.import_jobs[job_id] = {}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as source:
            source_path = Path(source.name)

        main.import_customers(job_id, source_path, ".xlsx", "new", "phone")

        job = main.import_jobs.pop(job_id)
        self.assertEqual(job["status"], "ready")
        self.assertEqual(
            [len(call.args[1]) for call in find_duplicates.call_args_list],
            [5_000, 1],
        )
        self.assertEqual(
            [len(call.args[1]) for call in create_batch.call_args_list],
            [5_000, 1],
        )

    def test_duplicate_lookup_batches_values_by_key(self):
        class Response:
            status_code = 200

            def __init__(self, data):
                self.data = data

            def json(self):
                return self.data

        class Session:
            def __init__(self):
                self.calls = []

            def get(self, url, **kwargs):
                self.calls.append(url)
                if "normalized_phone=in." in url:
                    return Response([{"id": 11, "normalized_phone": "9000000001"}])
                return Response([{"id": 12, "email": "aarav@example.com"}])

        session = Session()
        customers = [
            {"name": "Aarav", "phone": "9000000001", "email": "aarav@example.com"},
            {"name": "Priya", "phone": "9000000002", "email": "priya@example.com"},
        ]

        duplicate_ids = main._find_duplicates_batch(session, customers, ["phone", "email"])

        self.assertEqual(duplicate_ids, {0: 11})
        self.assertEqual(len(session.calls), 2)

    def test_duplicate_lookup_normalizes_phone_format(self):
        class Response:
            status_code = 200
            text = ""

            def json(self):
                return [{"id": 11, "normalized_phone": "919876543210"}]

        class Session:
            def __init__(self):
                self.url = ""

            def get(self, url, **kwargs):
                self.url = url
                return Response()

        session = Session()
        duplicate_ids = main._find_duplicates_batch(
            session,
            [{"name": "Aarav", "phone": "+91 98765-43210"}],
            ["phone"],
        )

        self.assertEqual(duplicate_ids, {0: 11})
        self.assertIn("normalized_phone=in.(919876543210)", session.url)

    def test_duplicate_rows_in_same_batch_are_removed(self):
        customers = [
            {"name": "First", "phone": "+91 98765-43210", "email": "first@example.com"},
            {"name": "Repeated phone", "phone": "919876543210", "email": "second@example.com"},
            {"name": "Repeated email", "phone": "9000000002", "email": "FIRST@EXAMPLE.COM"},
            {"name": "Unique", "phone": "9000000003", "email": "unique@example.com"},
        ]

        unique, duplicate_count = main._deduplicate_import_batch(
            customers, ["phone", "email"]
        )

        self.assertEqual([customer["name"] for customer in unique], ["First", "Unique"])
        self.assertEqual(duplicate_count, 2)

    def test_shared_phone_does_not_merge_customers_with_different_voter_ids(self):
        customers = [
            {
                "name": "Sreenivas Murthy",
                "phone": "9246842781",
                "voter_id_number": "GBZ0000001",
            },
            {
                "name": "Sanam Prema Latha",
                "phone": "9246842781",
                "voter_id_number": "GBZ7125826",
            },
        ]

        unique, duplicate_count = main._deduplicate_import_batch(
            customers, ["phone", "voter_id_number"]
        )

        self.assertEqual([customer["name"] for customer in unique], [
            "Sreenivas Murthy",
            "Sanam Prema Latha",
        ])
        self.assertEqual(duplicate_count, 0)

    def test_voter_id_lookup_takes_priority_over_shared_phone(self):
        class Response:
            status_code = 200
            text = ""

            def json(self):
                return [{"id": 42, "normalized_voter_id": "gbz7125826"}]

        class Session:
            def __init__(self):
                self.calls = []

            def get(self, url, **kwargs):
                self.calls.append(url)
                return Response()

        session = Session()
        duplicate_ids = main._find_duplicates_batch(
            session,
            [
                {
                    "name": "Sanam Prema Latha",
                    "phone": "9246842781",
                    "voter_id_number": "GBZ 7125826",
                }
            ],
            ["phone", "voter_id_number"],
        )

        self.assertEqual(duplicate_ids, {0: 42})
        self.assertEqual(len(session.calls), 1)
        self.assertIn("normalized_voter_id=in.(gbz7125826)", session.calls[0])

    @patch("backend.main._create_customers_batch")
    @patch("backend.main._find_duplicates_batch", return_value={})
    @patch("backend.main.iter_import_rows")
    def test_import_skips_normalized_duplicates_in_the_same_batch(
        self, iter_rows, _find_duplicates, create_batch
    ):
        iter_rows.return_value = iter(
            [
                {"name": "First", "phone": "+91 98765-43210"},
                {"name": "Repeated", "phone": "919876543210"},
            ]
        )
        job_id = "same-batch-duplicate-test"
        main.import_jobs[job_id] = {}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as source:
            source_path = Path(source.name)

        main.import_customers(job_id, source_path, ".xlsx", "new", "phone")

        job = main.import_jobs.pop(job_id)
        self.assertEqual(job["status"], "ready")
        self.assertEqual(job["created"], 1)
        self.assertEqual(job["skipped"], 1)
        create_batch.assert_called_once()
        self.assertEqual(len(create_batch.call_args.args[1]), 1)

    def test_customer_creation_sends_one_batch(self):
        class Response:
            status_code = 201

            def __init__(self):
                self.text = ""

        class Session:
            def __init__(self):
                self.payload = None
                self.headers = None

            def post(self, url, **kwargs):
                self.payload = kwargs["json"]
                self.headers = kwargs["headers"]
                return Response()

        session = Session()
        customers = [{"name": "Aarav", "phone": "9000000001"}]

        main._create_customers_batch(session, customers)

        self.assertEqual(session.payload, customers)
        self.assertEqual(session.headers["Prefer"], "return=minimal")

    def test_customer_creation_splits_large_chunks_into_safe_write_batches(self):
        class Response:
            status_code = 201
            text = ""

        class Session:
            def __init__(self):
                self.payloads = []

            def post(self, url, **kwargs):
                self.payloads.append(kwargs["json"])
                return Response()

        session = Session()
        customers = [
            {"name": f"Customer {index}", "phone": str(index)}
            for index in range(2_500)
        ]

        main._create_customers_batch(session, customers)

        self.assertEqual(
            [len(payload) for payload in session.payloads],
            [500, 500, 500, 500, 500],
        )

    def test_statement_timeout_splits_only_the_failed_write_batch(self):
        class Response:
            text = '{"code":"57014","message":"canceling statement due to statement timeout"}'

            def __init__(self, timed_out):
                self.status_code = 500 if timed_out else 201
                self.timed_out = timed_out

            def json(self):
                return {"code": "57014"} if self.timed_out else {}

        class Session:
            def __init__(self):
                self.batch_sizes = []

            def post(self, url, **kwargs):
                batch_size = len(kwargs["json"])
                self.batch_sizes.append(batch_size)
                return Response(batch_size > 250)

        session = Session()
        customers = [
            {"name": f"Customer {index}", "phone": str(index)}
            for index in range(500)
        ]

        main._create_customers_batch(session, customers)

        self.assertEqual(session.batch_sizes, [500, 250, 250])

    def test_single_customer_statement_timeout_is_not_retried_recursively(self):
        class Response:
            status_code = 500
            text = '{"code":"57014","message":"statement timeout"}'

            def json(self):
                return {"code": "57014"}

        class Session:
            def __init__(self):
                self.calls = 0

            def post(self, url, **kwargs):
                self.calls += 1
                return Response()

        session = Session()

        with self.assertRaisesRegex(ValueError, "statement timeout"):
            main._create_customers_batch(
                session,
                [{"name": "Aarav", "phone": "9000000001"}],
            )

        self.assertEqual(session.calls, 1)

    def test_timed_out_customer_update_is_retried_once(self):
        class Response:
            text = '{"code":"57014","message":"statement timeout"}'

            def __init__(self, status_code):
                self.status_code = status_code

            def json(self):
                return {"code": "57014"} if self.status_code == 500 else {}

        class Session:
            def __init__(self):
                self.calls = []

            def patch(self, url, **kwargs):
                self.calls.append(kwargs)
                return Response(500 if len(self.calls) == 1 else 204)

        session = Session()

        main._update_import_customer(
            session,
            42,
            {"name": "Aarav", "phone": "9000000001"},
        )

        self.assertEqual(len(session.calls), 2)
        self.assertTrue(
            all(call["headers"]["Prefer"] == "return=minimal" for call in session.calls)
        )

    def test_non_timeout_write_error_is_not_retried(self):
        class Response:
            status_code = 409
            text = '{"code":"23505","message":"duplicate key"}'

            def json(self):
                return {"code": "23505"}

        class Session:
            def __init__(self):
                self.calls = 0

            def post(self, url, **kwargs):
                self.calls += 1
                return Response()

        session = Session()

        with self.assertRaisesRegex(ValueError, "duplicate key"):
            main._create_customers_batch(
                session,
                [{"name": "Aarav", "phone": "9000000001"}],
            )

        self.assertEqual(session.calls, 1)

    @patch(
        "backend.main._create_customers_batch",
        side_effect=ValueError(
            '{"code":"57014","message":"canceling statement due to statement timeout"}'
        ),
    )
    @patch("backend.main._find_duplicates_batch", return_value={})
    @patch("backend.main.iter_import_rows")
    def test_import_failure_reports_spreadsheet_row_range(
        self, iter_rows, _find_duplicates, _create_batch
    ):
        iter_rows.return_value = iter(
            [
                {"name": "Aarav", "phone": "9000000001"},
                {"name": "Priya", "phone": "9000000002"},
            ]
        )
        job_id = "timeout-row-range-test"
        main.import_jobs[job_id] = {}
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as source:
            source_path = Path(source.name)

        main.import_customers(job_id, source_path, ".xlsx", "new", "phone")

        job = main.import_jobs.pop(job_id)
        self.assertEqual(job["status"], "failed")
        self.assertIn("Spreadsheet data rows 1-2", job["message"])
        self.assertIn('"code":"57014"', job["message"])

    def test_vercel_preview_origins_are_allowed_by_cors(self):
        with TestClient(main.app) as client:
            response = client.options(
                "/imports/customers",
                headers={
                    "Origin": "https://onecroredatapluse-git-main-dataonecrore.vercel.app",
                    "Access-Control-Request-Method": "POST",
                    "Access-Control-Request-Headers": "authorization",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["access-control-allow-origin"],
            "https://onecroredatapluse-git-main-dataonecrore.vercel.app",
        )

    def test_local_vite_preview_ports_are_allowed_by_cors(self):
        with TestClient(main.app) as client:
            response = client.options(
                "/imports/customers",
                headers={
                    "Origin": "http://localhost:5174",
                    "Access-Control-Request-Method": "POST",
                    "Access-Control-Request-Headers": "authorization",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["access-control-allow-origin"],
            "http://localhost:5174",
        )

    def test_confirmed_headers_map_to_customer_fields(self):
        row = main.normalize_import_row(
            {
                "Customer Name": " Priya   Reddy ",
                "Customer Phone": 9000000002,
                "Customer Address": "45, Green Park Colony,   Kondapur",
                "City": "Hyderabad",
                "State": "Telangana",
                "PIN Code": 500084,
            }
        )

        self.assertEqual(row["name"], "Priya Reddy")
        self.assertEqual(row["phone"], "9000000002")
        self.assertEqual(
            row["address"],
            "45, Green Park Colony, Kondapur, Hyderabad, Telangana, 500084",
        )

    def test_misspelled_customer_address_header_keeps_full_address(self):
        row = main.normalize_import_row(
            {
                "Customer Name": "Devadas M S",
                "Customer Phone": "9014168580",
                "Customer Adress": "12, Lake View Road",
                "City": "Karkala",
                "State": "Karnataka",
                "PIN Code": 500047,
            }
        )

        self.assertEqual(
            row["address"],
            "12, Lake View Road, Karkala, Karnataka, 500047",
        )

    def test_split_address_headers_are_combined_for_display(self):
        row = main.normalize_import_row(
            {
                "Name": "Tumma Govardhan",
                "Phone Number": "919949024248",
                "House No": "26/5/4/3/1",
                "Street": "BALRAM NAGAR",
                "Village": "BALRAMNAGAR",
                "Post office": "MALKajgiri",
                "Pincode": 500047,
            }
        )

        self.assertEqual(
            row["address"],
            "26/5/4/3/1, BALRAM NAGAR, BALRAMNAGAR, MALKajgiri, 500047",
        )

    def test_xlsx_with_confirmed_headers_is_accepted(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sample_customer_dataset.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Customer Name",
                    "Customer Phone",
                    "Customer Address",
                    "City",
                    "State",
                    "PIN Code",
                ]
            )
            sheet.append(
                [
                    "Aarav Sharma",
                    9000000001,
                    "12, Lake View Road, Banjara Hills",
                    "Hyderabad",
                    "Telangana",
                    500034,
                ]
            )
            workbook.save(path)

            rows = main.read_import_rows(path, ".xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "Aarav Sharma")
        self.assertEqual(rows[0]["phone"], "9000000001")
        self.assertEqual(
            rows[0]["address"],
            "12, Lake View Road, Banjara Hills, Hyderabad, Telangana, 500034",
        )

    def test_xlsx_rows_can_be_iterated_without_eager_loading(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "streamed_customers.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Customer Name", "Customer Phone"])
            sheet.append(["Aarav Sharma", 9000000001])
            workbook.save(path)

            rows = main.iter_import_rows(path, ".xlsx")
            first_row = next(rows)
            rows.close()

        self.assertEqual(first_row["name"], "Aarav Sharma")
        self.assertEqual(first_row["phone"], "9000000001")

    def test_xlsx_with_display_headers_is_accepted(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "display_headers.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Phone Number", "Address"])
            sheet.append(["Aarav Sharma", 9000000001, "12, Lake View Road"])
            workbook.save(path)

            rows = main.read_import_rows(path, ".xlsx")

        self.assertEqual(rows[0]["name"], "Aarav Sharma")
        self.assertEqual(rows[0]["phone"], "9000000001")
        self.assertEqual(rows[0]["address"], "12, Lake View Road")

    def test_common_phone_header_variants_are_accepted(self):
        self.assertEqual(
            main.normalize_import_row(
                {"Full Name": "Aarav Sharma", "Phone No.": 9000000001}
            ),
            {"name": "Aarav Sharma", "phone": "9000000001"},
        )

    def test_customer_spreadsheet_headers_map_to_database_fields(self):
        row = main.normalize_import_row(
            {
                "Name": "Aarav Sharma",
                "Phone Number": 9000000001,
                "Alternate Number": 9000000002,
                "Relationship Type": "Father",
                "Relationship Name": "Ramesh Sharma",
                "Voter ID Number": "ABC1234567",
                "Aadhar Card Number": 123456789012,
                "Email": "aarav@example.com",
                "Adress": "12, Lake View Road",
            }
        )

        self.assertEqual(
            row,
            {
                "name": "Aarav Sharma",
                "phone": "9000000001",
                "whatsapp_phone": "9000000002",
                "relationship_type": "Father",
                "relationship_name": "Ramesh Sharma",
                "voter_id_number": "ABC1234567",
                "aadhar_card_number": "123456789012",
                "email": "aarav@example.com",
                "address": "12, Lake View Road",
            },
        )

    def test_misspelled_relationship_name_header_is_preserved(self):
        row = main.normalize_import_row(
            {
                "Name": "Suresh Jetta",
                "Phone Number": 9502262279,
                "Alternate Number": None,
                "Relationship type": "FTHR",
                "Rationaship Name": "Anjaneyulu Jetta",
                "Voter Id": "NYY1219724",
                "House No": "2-46",
                "Village": "BONDUGULA",
                "Post office": "BONDUGULA",
                "Pincode": 508105,
            }
        )

        self.assertEqual(row["relationship_name"], "Anjaneyulu Jetta")

    def test_customer_phone_header_is_optional(self):
        main.validate_import_headers(["Customer Name", "Customer Address"])

    def test_missing_customer_name_has_clear_error(self):
        with self.assertRaisesRegex(ValueError, "Name column"):
            main.validate_import_headers(["Phone Number", "Customer Address"])


if __name__ == "__main__":
    unittest.main()
